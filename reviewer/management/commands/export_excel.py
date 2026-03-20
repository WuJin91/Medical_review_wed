# endoscopy-project/reviewer/management/commands/export_excel.py

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment
from django.core.management.base import BaseCommand, CommandError
from django.db.models import Q
from reviewer.models import EndoscopyImage, ImageBatch, Annotation

def calculate_iou(boxA, boxB):
    """計算兩個標註框的 IoU (Intersection over Union)"""
    if boxA is None or boxB is None:
        return 0.0

    # 將中心點、寬高格式轉換為 (x1, y1, x2, y2) 格式
    boxA_x1 = boxA['x_center'] - boxA['width'] / 2
    boxA_y1 = boxA['y_center'] - boxA['height'] / 2
    boxA_x2 = boxA['x_center'] + boxA['width'] / 2
    boxA_y2 = boxA['y_center'] + boxA['height'] / 2

    boxB_x1 = boxB['x_center'] - boxB['width'] / 2
    boxB_y1 = boxB['y_center'] - boxB['height'] / 2
    boxB_x2 = boxB['x_center'] + boxB['width'] / 2
    boxB_y2 = boxB['y_center'] + boxB['height'] / 2

    # 計算交集矩形的座標
    inter_x1 = max(boxA_x1, boxB_x1)
    inter_y1 = max(boxA_y1, boxB_y1)
    inter_x2 = min(boxA_x2, boxB_x2)
    inter_y2 = min(boxA_y2, boxB_y2)

    # 計算交集面積
    inter_area = max(0, inter_x2 - inter_x1) * max(0, inter_y2 - inter_y1)
    if inter_area == 0:
        return 0.0

    # 計算各自面積與聯集面積
    boxA_area = boxA['width'] * boxA['height']
    boxB_area = boxB['width'] * boxB['height']
    union_area = boxA_area + boxB_area - inter_area

    return inter_area / union_area

class Command(BaseCommand):
    help = '匯出詳盡的標註審核紀錄與 AI 模型效能分析 Excel 報告'

    def add_arguments(self, parser):
        parser.add_argument('output_file', type=str, help='指定匯出的 Excel 檔案的完整路徑與檔名 (e.g., ~/report.xlsx)')
        parser.add_argument('--batch-name', type=str, required=False, help='(可選) 只分析特定批次的資料')
        parser.add_argument('--iou-threshold', type=float, default=0.5, help='(可選) 用於判斷 TP/FP 的 IoU 閾值，預設為 0.5')

    def handle(self, *args, **options):
        output_file = Path(options['output_file'])
        batch_name = options['batch_name']
        iou_threshold = options['iou_threshold']

        if output_file.suffix.lower() != '.xlsx':
            raise CommandError("錯誤：輸出檔案必須是 .xlsx 格式。")

        # --- 1. 查詢基礎資料 ---
        all_annotations = Annotation.objects.select_related('image', 'image__batch').order_by('image__original_image__name', 'id')
        if batch_name:
            all_annotations = all_annotations.filter(image__batch__name=batch_name)

        if not all_annotations.exists():
            self.stdout.write(self.style.WARNING("找不到任何標註資料可供分析。"))
            return

        # --- 2. 建立 Excel 並寫入 Raw Data 工作表 ---
        workbook = openpyxl.Workbook()
        raw_sheet = workbook.active
        raw_sheet.title = "Raw_Annotation_Audit"

        # 寫入表頭
        headers = [
            'ImageName', 'BatchName', 'ImageType', 'DoctorFinalDiagnosis', 'AnnotationID',
            'ClassLabel', 'SourceType', 'IsDeletedByDoctor', 'ModelBox_Coordinates',
            'DoctorBox_Coordinates', 'IoU_vs_DoctorBox'
        ]
        raw_sheet.append(headers)

        # 彙整圖片對應的醫師標註框
        image_doctor_boxes = {}
        for ann in all_annotations:
            if ann.doctor_box and not ann.is_deleted:
                image_id = ann.image_id
                if image_id not in image_doctor_boxes:
                    image_doctor_boxes[image_id] = []
                image_doctor_boxes[image_id].append(ann)

        for ann in all_annotations:
            # 計算 IoU
            best_iou = 0.0
            if ann.model_box:
                doc_boxes_for_img = image_doctor_boxes.get(ann.image_id, [])
                for doc_ann in doc_boxes_for_img:
                    # 只有同類別的才計算IoU
                    if doc_ann.class_label == ann.class_label:
                        iou = calculate_iou(ann.model_box, doc_ann.doctor_box)
                        if iou > best_iou:
                            best_iou = iou
            
            row = [
                Path(ann.image.original_image.name).name, ann.image.batch.name, ann.image.get_image_type_display(),
                ann.image.get_doctor_diagnosis_display(), ann.id, ann.get_class_label_display(),
                ann.get_source_type_display(), ann.is_deleted,
                f"{ann.model_box['x_center']},{ann.model_box['y_center']},{ann.model_box['width']},{ann.model_box['height']}" if ann.model_box else "N/A",
                f"{ann.doctor_box['x_center']},{ann.doctor_box['y_center']},{ann.doctor_box['width']},{ann.doctor_box['height']}" if ann.doctor_box else "N/A",
                f"{best_iou:.4f}" if best_iou > 0 else "N/A"
            ]
            raw_sheet.append(row)

        self.stdout.write("工作表一：詳細標註審核紀錄... 完成。")

        # --- 3. 計算效能指標 ---
        stats = {
            'total_model_boxes': {'polyp': 0, 'tumor': 0},
            'total_doctor_boxes': {'polyp': 0, 'tumor': 0},
            'tp': {'polyp': 0, 'tumor': 0},
            'fp': {'polyp': 0, 'tumor': 0},
            'fn': {'polyp': 0, 'tumor': 0},
            'class_mismatch': {'polyp_as_tumor': 0, 'tumor_as_polyp': 0},
            'tp_ious': []
        }
        
        images_to_process = EndoscopyImage.objects.filter(id__in=all_annotations.values_list('image_id', flat=True).distinct())

        for image in images_to_process:
            model_anns = list(all_annotations.filter(image=image, source_type=Annotation.SourceTypes.MODEL))
            doctor_anns = list(all_annotations.filter(image=image, source_type=Annotation.SourceTypes.DOCTOR)) \
                        + list(all_annotations.filter(image=image, source_type=Annotation.SourceTypes.MODEL, doctor_box__isnull=False, is_deleted=False))
            
            # 去重
            doctor_anns = list({ann.id: ann for ann in doctor_anns}.values())

            # 統計總數
            for ann in model_anns: stats['total_model_boxes']['polyp' if ann.class_label == 0 else 'tumor'] += 1
            for ann in doctor_anns: stats['total_doctor_boxes']['polyp' if ann.class_label == 0 else 'tumor'] += 1

            matches = []
            
            # 尋找所有可能的匹配
            for m_ann in model_anns:
                for d_ann in doctor_anns:
                    iou = calculate_iou(m_ann.model_box, d_ann.doctor_box)
                    if iou >= iou_threshold:
                        matches.append({'model': m_ann, 'doctor': d_ann, 'iou': iou})
            
            # 貪婪匹配演算法
            matched_model_ids = set()
            matched_doctor_ids = set()

            # 依 IoU 降序排序，優先處理最匹配的
            matches.sort(key=lambda x: x['iou'], reverse=True)
            
            for match in matches:
                m_id, d_id = match['model'].id, match['doctor'].id
                if m_id not in matched_model_ids and d_id not in matched_doctor_ids:
                    matched_model_ids.add(m_id)
                    matched_doctor_ids.add(d_id)
                    
                    label_key = 'polyp' if match['model'].class_label == 0 else 'tumor'
                    if match['model'].class_label == match['doctor'].class_label:
                        stats['tp'][label_key] += 1
                        stats['tp_ious'].append(match['iou'])
                    else: # 類別錯誤
                        stats['fp'][label_key] += 1
                        if label_key == 'polyp':
                            stats['class_mismatch']['polyp_as_tumor'] += 1
                        else:
                            stats['class_mismatch']['tumor_as_polyp'] += 1
            
            # 計算 FP 和 FN
            for m_ann in model_anns:
                if m_ann.id not in matched_model_ids:
                    stats['fp']['polyp' if m_ann.class_label == 0 else 'tumor'] += 1
            
            for d_ann in doctor_anns:
                if d_ann.id not in matched_doctor_ids:
                    stats['fn']['polyp' if d_ann.class_label == 0 else 'tumor'] += 1

        self.stdout.write("效能指標計算... 完成。")

        # --- 4. 寫入 Summary 工作表 ---
        summary_sheet = workbook.create_sheet(title="Performance_Summary")
        
        # 輔助函式，用於寫入標題和數據
        def write_block(sheet, start_row, title, data):
            sheet.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=14)
            r = start_row + 1
            for key, value in data.items():
                sheet.cell(row=r, column=1, value=key)
                sheet.cell(row=r, column=2, value=value)
                r += 1
            return r + 1 # 回傳下一個區塊的起始行

        # 計算指標
        def calc_metrics(tp, fp, fn):
            precision = tp / (tp + fp) if (tp + fp) > 0 else 0
            recall = tp / (tp + fn) if (tp + fn) > 0 else 0
            f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
            return f"{precision:.2%}", f"{recall:.2%}", f"{f1:.4f}"

        polyp_p, polyp_r, polyp_f1 = calc_metrics(stats['tp']['polyp'], stats['fp']['polyp'], stats['fn']['polyp'])
        tumor_p, tumor_r, tumor_f1 = calc_metrics(stats['tp']['tumor'], stats['fp']['tumor'], stats['fn']['tumor'])
        avg_iou = sum(stats['tp_ious']) / len(stats['tp_ious']) if stats['tp_ious'] else 0

        # 寫入數據
        row_cursor = 1
        summary_data = {
            "分析批次": batch_name or "全部",
            "IoU 閾值": iou_threshold,
            "模型標註總數 (瘜肉)": stats['total_model_boxes']['polyp'],
            "模型標註總數 (腫瘤)": stats['total_model_boxes']['tumor'],
            "醫師標註總數 (瘜肉)": stats['total_doctor_boxes']['polyp'],
            "醫師標註總數 (腫瘤)": stats['total_doctor_boxes']['tumor'],
            "匹配框平均 IoU": f"{avg_iou:.4f}",
        }
        row_cursor = write_block(summary_sheet, row_cursor, "總體統計", summary_data)

        # 混淆矩陣
        summary_sheet.cell(row=row_cursor, column=1, value="二維混淆矩陣").font = Font(bold=True, size=14)
        conf_headers = ["(實際) / (預測)", "預測為瘜肉", "預測為腫瘤", "未檢出 (FN)"]
        summary_sheet.append(conf_headers)
        summary_sheet.append([
            "實際為瘜肉", stats['tp']['polyp'], stats['class_mismatch']['tumor_as_polyp'], stats['fn']['polyp']
        ])
        summary_sheet.append([
            "實際為腫瘤", stats['class_mismatch']['polyp_as_tumor'], stats['tp']['tumor'], stats['fn']['tumor']
        ])
        summary_sheet.append([
            "背景 (FP)", stats['fp']['polyp'] - stats['class_mismatch']['tumor_as_polyp'], stats['fp']['tumor'] - stats['class_mismatch']['polyp_as_tumor'], "N/A"
        ])
        row_cursor += 5

        # 各類別指標
        kpi_data = {
            "瘜肉 精確率 (Precision)": polyp_p,
            "瘜肉 召回率 (Recall)": polyp_r,
            "瘜肉 F1-Score": polyp_f1,
            "---": "---",
            "腫瘤 精確率 (Precision)": tumor_p,
            "腫瘤 召回率 (Recall)": tumor_r,
            "腫瘤 F1-Score": tumor_f1,
        }
        write_block(summary_sheet, row_cursor, "各類別關鍵績效指標 (KPIs)", kpi_data)

        self.stdout.write("工作表二：模型效能評估摘要... 完成。")

        # --- 5. 儲存檔案 ---
        try:
            workbook.save(output_file)
            self.stdout.write(self.style.SUCCESS(f"\n報告成功匯出至：{output_file}"))
        except Exception as e:
            raise CommandError(f"儲存 Excel 檔案時發生錯誤: {e}")